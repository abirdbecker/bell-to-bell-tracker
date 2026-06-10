#!/usr/bin/env python3
"""
One-time migration: PA Unplugged BELL TO BELL xlsx  ->  public/data/schools.json

After this runs, schools.json is the SOURCE OF TRUTH. The spreadsheet is retired.
All future edits happen through the /admin page (which commits schools.json back to
the repo via the GitHub API) or the public submission form (which appends to
pending.json for review).

Run:  python3 scripts/migrate-from-xlsx.py "/path/to/PA Unplugged BELL TO BELL School Policy Tracker.xlsx"
"""
import json, re, sys, hashlib
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path.home() / "Downloads" / "PA Unplugged BELL TO BELL School Policy Tracker.xlsx"

# Town -> county. Filled in by hand for the ~31 PA schools currently tracked.
# (Most map work uses county centroids, so we only need county-level accuracy.)
TOWN_COUNTY = {
    "Bradford": "McKean",
    "Carnegie": "Allegheny",
    "Clairton": "Allegheny",
    "Hershey": "Dauphin",
    "Harrisburg": "Dauphin",
    "Erie": "Erie",
    "McKeesport": "Allegheny",
    "Millcreek Township": "Erie",
    "Montoursville": "Lycoming",
    "Muhlenberg": "Berks",
    "Pittsburgh": "Allegheny",
    "Pittsburgh (Jackson Township)": "Butler",
    "Radnor": "Delaware",
    "New Eagle": "Washington",
    "Scranton": "Lackawanna",
    "Slippery Rock": "Butler",
    "Hookstown": "Beaver",
    "Mckees Rocks and Stowe": "Allegheny",
    "Washington": "Washington",
    "Wilkes-Barre": "Luzerne",
    "Midland": "Beaver",
    "Mt. Lebanon": "Allegheny",
    "Bryn Mawr": "Montgomery",
    "Rosemont": "Montgomery",
    "Springfield": "Delaware",
    "Philadelphia": "Philadelphia",
}
# Schools whose Town/City cell was blank in the sheet — assign county directly.
NAME_COUNTY = {
    "Brownsville Area School District": "Fayette",
}

# County centroids [lng, lat] — mirror of src/data/paCountyCentroids.js
def load_centroids():
    txt = (ROOT / "src" / "data" / "paCountyCentroids.js").read_text()
    cents = {}
    for m in re.finditer(r"name:\s*'([^']+)',\s*lng:\s*(-?[\d.]+),\s*lat:\s*(-?[\d.]+)", txt):
        cents[m.group(1)] = (float(m.group(2)), float(m.group(3)))
    return cents

CENTROIDS = load_centroids()

# Normalize the free-text "Phone Storage Location" into a filterable category.
def categorize_storage(raw):
    if not raw:
        return ("unknown", "Not specified")
    s = raw.strip()
    low = s.lower()
    # Mixed = different rules by grade band / building, or "varies"
    mixed_signals = ["k-6", "7-12", "lower school", "upper school", "varies", "\n"]
    has_mixed = any(sig in low for sig in mixed_signals)
    if "yondr" in low and not has_mixed:
        return ("yondr", "Yondr pouches")
    if has_mixed:
        return ("mixed", "Varies by grade/school")
    if "locker" in low:
        return ("lockers", "Stored in lockers")
    if "off and away" in low or "off & away" in low:
        return ("off_away", "Off and away")
    if "staff" in low or "collect" in low or "homeroom" in low:
        return ("staff", "Collected by staff")
    return ("other", s[:40])

def parse_students(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    digits = re.sub(r"[^\d]", "", str(v))
    return int(digits) if digits else None

def parse_year(effective):
    if effective is None:
        return None
    m = re.search(r"(19|20)\d{2}", str(effective))
    return int(m.group(0)) if m else None

def sector_from_type(t):
    if not t:
        return "public"
    t = t.strip().lower()
    if "charter" in t:
        return "charter"
    if "catholic" in t:
        return "catholic"
    if "private" in t:
        return "private"
    return "public"

def slugify(name):
    base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return base[:60]

def coords_for(county, used):
    """County centroid with deterministic jitter so co-located schools don't stack."""
    if county not in CENTROIDS:
        return (None, None)
    lng, lat = CENTROIDS[county]
    n = used.get(county, 0)
    used[county] = n + 1
    if n:
        # spiral the nth pin out from the centroid
        ang = n * 2.39996
        rad = 0.06 + 0.03 * n
        lng += rad * __import__("math").cos(ang)
        lat += rad * __import__("math").sin(ang)
    return (round(lng, 4), round(lat, 4))

def get_links(ws, row, source_col, policy_col):
    links = []
    for col, kind, default_label in [(source_col, "article", "Article"), (policy_col, "policy", "Policy")]:
        cell = ws.cell(row=row, column=col)
        if cell.hyperlink and cell.hyperlink.target:
            label = (str(cell.value).strip() if cell.value else default_label) or default_label
            links.append({"kind": kind, "label": label, "url": cell.hyperlink.target})
    return links

def build(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    used = {}
    schools = []

    sheets = [
        # (sheet, header_row, name_col, storage, town, level, type, students, effective, notes, source_col, policy_col)
        ("PA Public Schools", 2, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12),
        ("PA PrivateCharter", 1, 1, 2, 3, 5, 6, 7, 8, 9, 10, 11),
    ]
    for (sheet, hdr, c_name, c_store, c_town, c_level, c_type, c_stud, c_eff, c_notes, c_src, c_pol) in sheets:
        ws = wb[sheet]
        for r in range(hdr + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=c_name).value
            if not name:
                continue
            name = str(name).strip()
            town = ws.cell(row=r, column=c_town).value
            town = str(town).strip() if town else None
            county = NAME_COUNTY.get(name) or (TOWN_COUNTY.get(town) if town else None)
            lng, lat = coords_for(county, used) if county else (None, None)
            cat, label = categorize_storage(ws.cell(row=r, column=c_store).value)
            notes = ws.cell(row=r, column=c_notes).value
            contact = ws.cell(row=r, column=c_eff + 4).value if False else ws.cell(row=r, column=c_pol + 1).value
            schools.append({
                "id": slugify(name),
                "name": name,
                "sector": sector_from_type(ws.cell(row=r, column=c_type).value),
                "typeRaw": (str(ws.cell(row=r, column=c_type).value).strip()
                            if ws.cell(row=r, column=c_type).value else None),
                "town": town,
                "county": county,
                "lng": lng,
                "lat": lat,
                "level": (str(ws.cell(row=r, column=c_level).value).strip()
                          if ws.cell(row=r, column=c_level).value else None),
                "students": parse_students(ws.cell(row=r, column=c_stud).value),
                "storage": {"category": cat, "label": label,
                            "raw": (str(ws.cell(row=r, column=c_store).value).strip()
                                    if ws.cell(row=r, column=c_store).value else None)},
                "effective": (str(ws.cell(row=r, column=c_eff).value).strip()
                              if ws.cell(row=r, column=c_eff).value else None),
                "year": parse_year(ws.cell(row=r, column=c_eff).value),
                "notes": (str(notes).strip() if notes else None),
                "links": get_links(ws, r, c_src, c_pol),
                "contact": (str(contact).strip() if contact else None),
                "status": "published",
            })

    schools.sort(key=lambda s: s["name"])
    out = {
        "updatedAt": __import__("datetime").date.today().isoformat(),
        "schools": schools,
    }
    dest = ROOT / "public" / "data" / "schools.json"
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n")
    # initialize empty review queue
    pend = ROOT / "public" / "data" / "pending.json"
    if not pend.exists():
        pend.write_text(json.dumps({"pending": []}, indent=2) + "\n")

    no_coords = [s["name"] for s in schools if s["lat"] is None]
    print(f"Wrote {len(schools)} schools -> {dest}")
    print(f"  by sector: " + ", ".join(f"{k}={sum(1 for s in schools if s['sector']==k)}"
                                        for k in ["public", "charter", "catholic", "private"]))
    if no_coords:
        print(f"  WARNING: no map coords for: {no_coords}")

if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    build(path)
